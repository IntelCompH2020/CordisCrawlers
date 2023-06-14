
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

# def load_xlsx_dato(my_path):
#     return pd.read_excel(my_path).fillna('N/A').to_dict('records')

def load_data_from_excel(xlsx_fpath):
    wb = load_workbook(xlsx_fpath)
    xl_data = {}
    for sheet_name in wb.sheetnames:
        temp_data = []
        sheet = wb[sheet_name]
        for row in tqdm(sheet.iter_rows()):
            temp_data.append([str(item.value).strip() if item.value else None for item in row])
        xl_data[sheet_name] = temp_data
    ######################
    return xl_data

###################################################################################################################
euroSciVoc = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/euroSciVoc.xlsx')['euroSciVoc']
euroSciVoc = pd.DataFrame(euroSciVoc[1:], columns =euroSciVoc[0])
euroSciVoc.to_parquet('euroSciVoc.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
legalBasis = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/legalBasis.xlsx')['legalBasis']
legalBasis = pd.DataFrame(legalBasis[1:], columns =legalBasis[0])
legalBasis.to_parquet('legalBasis.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
organization = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/organization.xlsx')['organization']
organization = pd.DataFrame(organization[1:], columns =organization[0])
organization.to_parquet('organization.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
projectDeliverables = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/projectDeliverables.xlsx')['projectDeliverables']
projectDeliverables = pd.DataFrame(projectDeliverables[1:], columns =projectDeliverables[0])
projectDeliverables.to_parquet('projectDeliverables.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
projectPublications = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/projectPublications.xlsx')['projectPublications']
projectPublications = pd.DataFrame(projectPublications[1:], columns =projectPublications[0])
projectPublications.to_parquet('projectPublications.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
project = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/project.xlsx')['project']
project = pd.DataFrame(project[1:], columns =project[0])
project.to_parquet('project.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
reportSummaries = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/reportSummaries.xlsx')['reportSummaries']
reportSummaries = pd.DataFrame(reportSummaries[1:], columns =reportSummaries[0])
reportSummaries.to_parquet('reportSummaries.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
topics = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/topics.xlsx')['topics']
topics = pd.DataFrame(topics[1:], columns =topics[0])
topics.to_parquet('topics.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
webItem = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/webItem.xlsx')['webItem']
webItem = pd.DataFrame(webItem[1:], columns =webItem[0])
webItem.to_parquet('webItem.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################
webLink = load_data_from_excel('/media/disk1/dpappas_data/cordis_dumps/h2020/xlsx/webLink.xlsx')['webLink']
webLink = pd.DataFrame(webLink[1:], columns =webLink[0])
webLink.to_parquet('webLink.parquet', engine='auto', compression='snappy', index=None, partition_cols=None)
###################################################################################################################

print(project['id'].unique().shape[0])
all_proj_ids = project['id'].unique().tolist()
topics   = pd.DataFrame(topics[1:], columns =topics[0])
print(topics['projectID'].unique().shape[0])
print(topics[topics['projectID'] == '871140']['topic'])

