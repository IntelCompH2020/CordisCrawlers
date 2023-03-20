
import os, json, zipfile
from tqdm import tqdm
from pprint import pprint
from openpyxl import load_workbook

def load_data_from_excel(xlsx_fpath):
    wb = load_workbook(xlsx_fpath)
    xl_data = {}
    for sheet_name in wb.sheetnames:
        temp_data = []
        sheet = wb[sheet_name]
        for row in tqdm(sheet.iter_rows()):
            temp_data.append([str(item.value).strip() if item.value else None for item in row])
        xl_data[sheet_name] = temp_data
    ######################
    return xl_data

def list2dict(l, key):
    ret = {}
    for item in l:
        ret.setdefault(item[key], []).append(item)
    return ret

directory   = './'

############################################################################################################

zip_path    = os.path.join(directory, 'cordisref-H2020programmes-xlsx.zip')
archive     = zipfile.ZipFile(zip_path, 'r')

programmes = []
for name2 in tqdm(archive.namelist()):
    print("Found internal internal file: " + name2)
    rows    = iter(load_data_from_excel(archive.open(name2, 'r')).values()).__next__()
    labels  = rows[0]
    for row in rows[1:]:
        programmes.append(dict(zip(labels, row)))

print(len(programmes))
programmes = list2dict(programmes, 'id')

############################################################################################################

# zip_path    = os.path.join(directory, 'cordisref-H2020topics-xlsx.zip')
# archive     = zipfile.ZipFile(zip_path, 'r')
#
# topics = []
# for name2 in tqdm(archive.namelist()):
#     print("Found internal internal file: " + name2)
#     rows    = iter(load_data_from_excel(archive.open(name2, 'r')).values()).__next__()
#     labels  = rows[0]
#     for row in rows[1:]:
#         topics.append(dict(zip(labels, row)))
#
# print(len(topics))
# topics = list2dict(topics, 'id')


############################################################################################################

zip_path    = os.path.join(directory, 'cordis-h2020reports-xlsx.zip')
archive     = zipfile.ZipFile(zip_path, 'r')

reports = []
for name2 in tqdm(archive.namelist()):
    print("Found internal internal file: " + name2)
    # rows    = iter(load_data_from_excel(archive.open('xlsx/projectPublications.xlsx', 'r')).values()).__next__()
    rows    = iter(load_data_from_excel(archive.open(name2, 'r')).values()).__next__()
    labels  = rows[0]
    for row in rows[1:]:
        reports.append(dict(zip(labels, row)))

print(len(reports))
reports = list2dict(reports, 'projectID')

############################################################################################################

zip_path    = os.path.join(directory, 'cordis-h2020projectDeliverables-xlsx.zip')
archive     = zipfile.ZipFile(zip_path, 'r')

deliverables = []
for name2 in tqdm(archive.namelist()):
    print("Found internal internal file: " + name2)
    # rows    = iter(load_data_from_excel(archive.open('xlsx/projectPublications.xlsx', 'r')).values()).__next__()
    rows    = iter(load_data_from_excel(archive.open(name2, 'r')).values()).__next__()
    labels  = rows[0]
    for row in rows[1:]:
        deliverables.append(dict(zip(labels, row)))

print(len(deliverables))
deliverables    = list2dict(deliverables, 'projectID')

############################################################################################################

zip_path    = os.path.join(directory, 'cordis-h2020projectPublications-xlsx.zip')
archive     = zipfile.ZipFile(zip_path, 'r')

pubs = []
for name2 in tqdm(archive.namelist()):
    # print("Found internal internal file: " + name2)
    # rows    = iter(load_data_from_excel(archive.open('xlsx/projectPublications.xlsx', 'r')).values()).__next__()
    rows    = iter(load_data_from_excel(archive.open(name2, 'r')).values()).__next__()
    labels  = rows[0]
    for row in rows[1:]:
        pubs.append(dict(zip(labels, row)))

print(len(pubs))
pubs        = list2dict(pubs, 'projectID')

############################################################################################################

zip_path    = os.path.join(directory, 'cordis-h2020projects-json.zip')

archive             = zipfile.ZipFile(zip_path, 'r')
for name2 in archive.namelist():
    print("Found internal internal file: " + name2)

project_data    = json.loads(str(archive.read('json/project.json').decode('utf-8')))
org_data        = json.loads(str(archive.read('json/organization.json').decode('utf-8')))
topics_data     = json.loads(str(archive.read('json/topics.json').decode('utf-8')))
legal_data      = json.loads(str(archive.read('json/legalBasis.json').decode('utf-8')))
euroscivoc_data = json.loads(str(archive.read('json/euroSciVoc.json').decode('utf-8')))
weblink_data    = json.loads(str(archive.read('json/webLink.json').decode('utf-8')))

org_data        = list2dict(org_data, 'projectID')
topics_data     = list2dict(topics_data, 'projectID')
legal_data      = list2dict(legal_data, 'projectID')
euroscivoc_data = list2dict(euroscivoc_data, 'projectID')
weblink_data    = list2dict(weblink_data, 'projectID')

############################################################################################################

exported_data   = []
for project_datum in tqdm(project_data):
    project_datum['orgs']           = org_data[project_datum['id']]             if project_datum['id'] in org_data else []
    project_datum['topics']         = topics_data[project_datum['id']]          if project_datum['id'] in topics_data else []
    project_datum['legal']          = legal_data[project_datum['id']]           if project_datum['id'] in legal_data else []
    project_datum['euroscivoc']     = euroscivoc_data[project_datum['id']]      if project_datum['id'] in euroscivoc_data else []
    project_datum['web']            = weblink_data[project_datum['id']]         if project_datum['id'] in weblink_data else []
    project_datum['publications']   = pubs[project_datum['id']]                 if project_datum['id'] in pubs else []
    project_datum['reports']        = reports[project_datum['id']]              if project_datum['id'] in reports else []
    project_datum['deliverables']   = deliverables[project_datum['id']]         if project_datum['id'] in deliverables else []
    project_datum['programmes']     = programmes[project_datum['legalBasis']]   if project_datum['legalBasis'] in programmes else []
    exported_data.append(project_datum)

json.dump(exported_data, open('h2020_project_metadata.json', 'w', encoding='utf-8'))
